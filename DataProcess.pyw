from tkinter import *
from tkinter import ttk 
from tkinter import messagebox
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook,Workbook
from statistics import mean, stdev
import os 
import sys

window = Tk()
window.title('数据处理 v3.0')
window.geometry('400x735')  

str_plate384filerequirement ="首先需要一个化合物Compound.xlsx文件，compound文件须包含MOLENAME/Plate location/cas/MolWt/Bioactivity/Formula/Status/Reference 列，然后要处理的文件包含Position384列,，该列格式为：Plate1 O9,处理后的数据保存在Position384Output.csv"
str_plate96filerequirement = "首先需要一个化合物Compound.xlsx文件，compound文件须包含MOLENAME/Plate location/cas/MolWt/Bioactivity/Formula/Status/Reference 列,然后要处理的文件包含Position列，该列格式为：1-A3,处理后的数据保存在PositionOutput.csv"
str_datascreeningfilerequirement = "文件表格的第一行必须要包含有Sequence/RawData/Position384/OriginalSequence四列，且RawData这一列需要是数字，处理后的数据保存在DataScreeningOutput.csv"
str_rawdataprocessfilerequirement = "384孔板号输出原始数据cvs文件"
str_multirawdataprocessfilerequirement = "该模板可处理多工作表原始数据，并在原始表格中生成运算结果；该模板为Activation%的计算公式;选择384孔板号原始数据xlsx文件;CV位置为C20;输出的文件为MultiRawDataOutput.csv"
str_extractrawdataprocessfilerequirement = "该模块是提取同一目录所有原始数据文件，汇集到同一个excel文件,文件夹内文件名的要求如下：每个文件名的格式均需统一，必须只能含有一个“-”，且在“-”之后一定要连接384板号，两位数形式，例如第1块板即为-01"

#sys.stderr = open('error.log','w')

########数据筛选处理#############
filename_list=[] 
comp_filename_list = []

def chosescreeningfile():
	filepath = filedialog.askopenfilenames(filetype=[("csv file","*.csv")])
	filename_list.clear()
	filename_list.append(filepath)
	str_list= filename_list[0][0].split('/')	
	lbfile.configure(text=str_list[-1])
def datascreening():
	
	Position384 = []
	Sequence = []
	RawData = []
	df = pd.read_csv(filename_list[0][0]) 
	for (index, row) in df.iterrows():
		if float( row.RawData) > float(threshold_entry.get()) : 
			Sequence.append(row.Sequence)
			RawData.append(row.RawData)
			Position384.append(df.Position384[df.OriginalSequence.tolist().index(row.Sequence)])

	outdf1 = pd.DataFrame({'Sequence':Sequence}) 
	outdf2 = pd.DataFrame({'RawData':RawData}) 
	outdf3 = pd.DataFrame({'Position384':Position384}) 
	outdf = pd.concat([outdf1,outdf2,outdf3],axis=1)
	outdf.index = outdf.index + 1
	outdf.to_csv("DataScreeningOutput.csv")
	messagebox.showinfo('提醒',"处理完成")
def datascreeningfilerequirement():
	messagebox.showinfo('文件要求',str_datascreeningfilerequirement)
	
ttk.Label(window,justify="left", text="SPR-384板数据筛选" ).grid(column=0, row=0)
ttk.Label(window, justify="left",text="文件名：").grid(column=0, row=1) 
lbfile =ttk.Label(window, text=" ")
lbfile.grid(column=1, row=1)
ttk.Button(window, text="文件说明", command=datascreeningfilerequirement).grid(column=0, row=2)
ttk.Button(window, text="选择文件", command=chosescreeningfile).grid(column=1, row=2)
ttk.Label(window,justify="left", text="阈值：" ).grid(column=2, row=2)
threshold_entry=ttk.Entry(window,width = 8 )
threshold_entry.grid(column=3, row=2) 
threshold_entry.insert(0,"30.0")
#threshold_entry.pack()
ttk.Button(window, text="处理文件", command=datascreening).grid(column=0, row=3)

###############384转96################################

def chose384file():
	filepath = filedialog.askopenfilenames(filetype=[("csv file","*.csv")])
	filename_list.clear()
	filename_list.append(filepath)
	
	str_list= filename_list[0][0].split('/')	
	lb384.configure(text=str_list[-1])
def chose384compoundfile():
	filepath = filedialog.askopenfilenames(filetype=[("xlsx file","*.xlsx")])
	comp_filename_list.clear()
	comp_filename_list.append(filepath) 
	
	str_list= comp_filename_list[0][0].split('/')	
	lb384comp.configure(text=str_list[-1])

def plate384to96():
	Alphabet_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
	plate96_list = []
	df = pd.read_csv(filename_list[0][0]) 

	#compound_df = pd.read_csv('Compound.csv')

	comp_wb = load_workbook(comp_filename_list[0][0])
	data = comp_wb.active.values
	columns = next(data)[0:]
	compound_df = pd.DataFrame(data, columns=columns) 

	MOLENAME_list = []
	MolWt_list = []
	cas_list =[]
	Bioactivity_list = []
	Formula_list = []
	Status_list = []
	Reference_list = []
	for (index, row) in df.iterrows():
		str_list=row.Position384.split()
		PlateNo = int(str_list[1])
		RowNo = Alphabet_list.index(str_list[2][0])
		ColumnNo = int(str_list[2][1:])
		
		plate96=(PlateNo-1)*4+1+(RowNo%2 *2 +(ColumnNo+1)%2)
		str_96= str(plate96)+"-"+Alphabet_list[int(RowNo/2)]+str(int((ColumnNo+1)/2))
		plate96_list.append(str_96)

		# look for compound 
		flag = False
		for (j, r) in compound_df.iterrows() : 
			if str_96 == (r['Plate location']) : 
				MOLENAME_list.append(r.MOLENAME)
				MolWt_list.append(r.MolWt)
				cas_list.append('#'+str(r.cas))
				Bioactivity_list.append(r.Bioactivity)
				Formula_list.append(r.Formula)
				Status_list.append(r.Status)
				Reference_list.append(r.Reference)
				flag=True
		
		if flag == False :
			MOLENAME_list.append("n/a")
			MolWt_list.append("n/a")
			cas_list.append("n/a")
			Bioactivity_list.append("n/a")
			Formula_list.append("n/a")
			Status_list.append("n/a")
			Reference_list.append("n/a")
			
	outdf = pd.DataFrame({'Position96':plate96_list}) 
	MOLENAMEdf = pd.DataFrame({'MOLENAME':MOLENAME_list}) 
	MolWtdf = pd.DataFrame({'MolWt':MolWt_list}) 
	casdf = pd.DataFrame({'cas':cas_list}) 
	Bioactivitydf = pd.DataFrame({'Bioactivity':Bioactivity_list})
	Formuladf = pd.DataFrame({'Formula':Formula_list})
	Statusdf = pd.DataFrame({'Status':Status_list})
	Referencedf = pd.DataFrame({'Reference':Reference_list})

	outdf['Position384']=df['Position384']
	outdf['MOLENAME']=MOLENAMEdf['MOLENAME']
	outdf['MolWt']=MolWtdf['MolWt']
	outdf['cas']=casdf['cas']
	outdf['Bioactivity']=Bioactivitydf['Bioactivity']
	outdf['Formula']=Formuladf['Formula']
	outdf['Status']=Statusdf['Status']
	outdf['Reference']=Referencedf['Reference']


	outdf.index = outdf.index + 1
	outdf.to_csv("Position384Output.csv")
	
	messagebox.showinfo('提醒',"处理完成") 
def plate384filerequirement() :
	messagebox.showinfo('提醒',str_plate384filerequirement)
	
ttk.Label(window, text="384板转96板" ).grid(column=0, row=6) 
ttk.Label(window, text="文件名：").grid(column=0, row=9) 
lb384 =ttk.Label(window, text=" ")
lb384.grid(column=1, row=9)

ttk.Label(window, text="文件名：").grid(column=0, row=8) 
lb384comp =ttk.Label(window, text=" ")
lb384comp.grid(column=1, row=8)

ttk.Button(window, text="选择文件", command=chose384file).grid(column=1, row=10)
ttk.Button(window, text="Compound", command=chose384compoundfile).grid(column=0, row=10)
ttk.Button(window, text="文件说明", command=plate384filerequirement).grid(column=0, row=11)
ttk.Button(window, text="处理文件", command=plate384to96).grid(column=1, row=11)


###################96板处理#####################
def chose96file():
	filepath = filedialog.askopenfilenames(filetype=[("csv file","*.csv")])
	filename_list.clear()
	filename_list.append(filepath) 
	
	str_list= filename_list[0][0].split('/')	
	lb96.configure(text=str_list[-1])
def chosecompoundfile():
	filepath = filedialog.askopenfilenames(filetype=[("xlsx file","*.xlsx")])
	comp_filename_list.clear()
	comp_filename_list.append(filepath) 
	
	str_list= comp_filename_list[0][0].split('/')	
	lbcomp.configure(text=str_list[-1])

def plate96process():
	
	df = pd.read_csv(filename_list[0][0])

	comp_wb = load_workbook(comp_filename_list[0][0])
	data = comp_wb.active.values
	columns = next(data)[0:]
	compound_df = pd.DataFrame(data, columns=columns) 
	#compound_df = pd.read_csv('Compound.csv')
	MOLENAME_list = []
	MolWt_list = []
	cas_list =[]
	Bioactivity_list = []
	Formula_list = []
	Status_list = []
	Reference_list = []

	for (index, row) in df.iterrows():
		flag=False
		for (j, r) in compound_df.iterrows() : 
			if row.Position == (r['Plate location']) : 
				MOLENAME_list.append(r.MOLENAME)
				MolWt_list.append(r.MolWt)
				cas_list.append('#'+str(r.cas))
				Bioactivity_list.append(r.Bioactivity)
				Formula_list.append(r.Formula)
				Status_list.append(r.Status)
				Reference_list.append(r.Reference)
				flag = True
			else :
				str1=row.Position
				strlist=str1.split('-')
				if(len(strlist[1]))==2 :
					str2=str1[:-1]+'0'+str1[-1:]
					if str2 == (r['Plate location']) : 
						MOLENAME_list.append(r.MOLENAME)
						MolWt_list.append(r.MolWt)
						cas_list.append('#'+str(r.cas))
						Bioactivity_list.append(r.Bioactivity)
						Formula_list.append(r.Formula)
						Status_list.append(r.Status)
						Reference_list.append(r.Reference)
						flag = True
				
		if flag ==False :
			MOLENAME_list.append("null")
			MolWt_list.append("null")
			cas_list.append("null")
			Bioactivity_list.append("null")
			Formula_list.append("null")
			Status_list.append("null")
			Reference_list.append("null")

	outdf = pd.DataFrame( ) 
	MOLENAMEdf = pd.DataFrame({'MOLENAME':MOLENAME_list}) 
	MolWtdf = pd.DataFrame({'MolWt':MolWt_list}) 
	casdf = pd.DataFrame({'cas':cas_list}) 
	Bioactivitydf = pd.DataFrame({'Bioactivity':Bioactivity_list})
	Formuladf = pd.DataFrame({'Formula':Formula_list})
	Statusdf = pd.DataFrame({'Status':Status_list})
	Referencedf = pd.DataFrame({'Reference':Reference_list})

	outdf['Position']=df['Position']
	outdf['MOLENAME']=MOLENAMEdf['MOLENAME']
	outdf['MolWt']=MolWtdf['MolWt']
	outdf['cas']=casdf['cas']
	outdf['Bioactivity']=Bioactivitydf['Bioactivity']
	outdf['Formula']=Formuladf['Formula']
	outdf['Status']=Statusdf['Status']
	outdf['Reference']=Referencedf['Reference']

	outdf.index = outdf.index + 1
	outdf.to_csv("PositionOutput.csv")
	messagebox.showinfo('提醒',"处理完成")

def plate96filerequirement() :
	messagebox.showinfo('提醒',str_plate96filerequirement)
	
ttk.Label(window, text="化合物信息输出" ).grid(column=0, row=12) 

ttk.Label(window, text="文件名：").grid(column=0, row=13) 
lbcomp =ttk.Label(window, text=" ")
lbcomp.grid(column=1, row=13)
ttk.Label(window, text="文件名：").grid(column=0, row=14) 
lb96 =ttk.Label(window, text=" ")
lb96.grid(column=1, row=14)
ttk.Button(window, text="文件说明", command=plate96filerequirement).grid(column=0, row=16)
ttk.Button(window, text="Compound", command=chosecompoundfile).grid(column=0, row=15)
ttk.Button(window, text="选择板号文件", command=chose96file).grid(column=1, row=15)
ttk.Button(window, text="处理文件", command=plate96process).grid(column=1, row=16)

###########################384 screening ##################################
def choserawdatafile():
	filepath = filedialog.askopenfilenames(filetype=[("csv file","*.csv")])
	filename_list.clear()
	filename_list.append(filepath) 
	
	str_list= filename_list[0][0].split('/')	
	lbrawdata.configure(text=str_list[-1])
def rawdataprocess():

	df = pd.read_csv(filename_list[0][0], header=None)
	if v.get() ==1:
		threshold= float(threshold384_entry.get())*float(df[2][20])
	else :
		threshold= float(threshold384_entry.get() )
	
	print("threshold = {}".format(threshold))
	Alphabet_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
	plate384_list = []
	plate96_list = []
	rawdata_list =[]
	for row in range(24,40) :
		for col in range(3,23) :
			#print("row = {} col={} value = {}  threshold = {}".format(row,col,float(df[col][row]),threshold))
			if float(df[col][row])> threshold :
				str_384 = "Plate "+df[0][0]+ " " +Alphabet_list[row-24] +str(col)
				plate96=(int(df[0][0])-1)*4+1+(row%2 *2 +(col+1)%2)
				
				str_96= str(plate96)+"-"+Alphabet_list[int((row-24)/2)]+str(int((col+1)/2))
				plate384_list.append(str_384)
				plate96_list.append(str_96)
				rawdata_list.append(float(df[col][row]))
				
	outdf = pd.DataFrame( ) 
	Plate384df = pd.DataFrame({'Position384':plate384_list}) 
	Plate96df = pd.DataFrame({'Position96':plate96_list}) 
	Performancedf = pd.DataFrame({'Performance':rawdata_list}) 

	outdf['Position384']=Plate384df['Position384']
	outdf['Position96']=Plate96df['Position96']
	outdf['Performance']=Performancedf['Performance'] 

	outdf.index = outdf.index + 1
	outdf.to_csv("RawDataOutput.csv")
	messagebox.showinfo('提醒',"处理完成")
	
def rawdataprocessfilerequirement() :
	messagebox.showinfo('提醒',str_rawdataprocessfilerequirement)
	
	
ttk.Label(window,  text="读取384 CSV 数据" ).grid(column=0, row=18) 
v= IntVar()
ttk.Radiobutton(window, text ="倍数",variable=v,value =1).grid(column=0,row= 19)
ttk.Radiobutton(window, text ="数值",variable=v,value =2).grid(column=1,row= 19)
v.set(1)

threshold384_entry=ttk.Entry(window,width = 8 )
threshold384_entry.grid(column=2, row=19) 
threshold384_entry.insert(0,"5")

ttk.Label(window, text="文件名：").grid(column=0, row=20) 
lbrawdata =ttk.Label(window, text=" ")
lbrawdata.grid(column=1, row=20)

ttk.Button(window, text="文件说明", command=rawdataprocessfilerequirement).grid(column=0, row=21)
ttk.Button(window, text="选择文件", command=choserawdatafile).grid(column=1, row=21)
ttk.Button(window, text="处理文件", command=rawdataprocess).grid(column=0, row=22)

###########################  multi 384 screening ##################################
def chosemultirawdatafile():
	filepath = filedialog.askopenfilenames(filetype=[("xlsx file","*.xlsx")])
	filename_list.clear()
	filename_list.append(filepath) 
	
	str_list= filename_list[0][0].split('/')	
	lbmultirawdata.configure(text=str_list[-1])

def multirawdataprocess():

	filepath = filename_list[0][0]
	wb = load_workbook(filepath)
	
	Alphabet_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
	plate384_list = []
	plate96_list = []
	rawdata_list =[]
	
	for sheetname in wb.sheetnames:
		sheet = wb[sheetname] 

		# write B18/B19/B20 cell value
		sheet['B18'].value = 'Average'
		sheet['B19'].value = 'SD'
		sheet['B20'].value = '%CV'
		sum = 0
		num_list = []
		for column in [3,24]:
			for row in range(2,18): 
				num_list.append(float(sheet.cell(row,column).value))
		sheet['C18'].value = mean(num_list)
		sheet['C19'].value = stdev(num_list)
		sheet['C20'].value = stdev(num_list)/mean(num_list)*100

		# input row 23 value
		sheet['A23'].value = 'Activation%'
		for column in range(2,26):
			sheet.cell(23,column).value = column-1
		for row in range(24,40):
			sheet.cell(row,1).value = sheet.cell(row-22,1).value
		for row in range(24,40):
			for column in range(3,25):
				sheet.cell(row,column).value = (float(sheet.cell(row-22,column).value)/mean(num_list) - 1)*100

		# df = pd.read_csv(filename_list[0][0], header=None)
		if mv.get() ==1:
			threshold= float(mthreshold384_entry.get())*(sheet.cell(20,3).value) # df[23][20]
		else :
			threshold= float(mthreshold384_entry.get() )
		
		for row in range(24,40) :
			for col in range(4,24) :
				if mv2.get() == 1: 
					if float(sheet.cell(row,col).value)> threshold :
						str_384 = "Plate "+str(sheet.cell(1,1).value)+ " " +Alphabet_list[row-24] +str(col-1)
						plate96=(int(sheet.cell(1,1).value)-1)*4+1+((row)%2 *2 +(col)%2)   
						str_96= str(plate96)+"-"+Alphabet_list[int((row-24)/2)]+str(int(col/2))
						plate384_list.append(str_384)
						plate96_list.append(str_96)
						rawdata_list.append(float(sheet.cell(row,col).value))
				else : 
					if float(sheet.cell(row,col).value)<= threshold :
						str_384 = "Plate "+str(sheet.cell(1,1).value)+ " " +Alphabet_list[row-24] +str(col-1) 
						plate96=(int(sheet.cell(1,1).value)-1)*4+1+((row)%2 *2 +(col)%2)
						str_96= str(plate96)+"-"+Alphabet_list[int((row-24)/2)]+str(int(col/2))
						plate384_list.append(str_384)
						plate96_list.append(str_96)
						rawdata_list.append(float(sheet.cell(row,col).value))
				
	outdf = pd.DataFrame( ) 
	Plate384df = pd.DataFrame({'Position384':plate384_list}) 
	Plate96df = pd.DataFrame({'Position96':plate96_list}) 
	Performancedf = pd.DataFrame({'Performance':rawdata_list}) 

	outdf['Position384']=Plate384df['Position384']
	outdf['Position96']=Plate96df['Position96']
	outdf['Performance']=Performancedf['Performance'] 

	outdf.index = outdf.index + 1
	if mv3.get() == 2 :
		filter384holeno(outdf).to_csv("MultiRawDataOutput.csv")
	else : 
		outdf.to_csv("MultiRawDataOutput.csv")

	wb.save(filepath)
	messagebox.showinfo('提醒',"处理完成")
def filter384holeno(df):
	Alphabet_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
	df = pd.read_csv('MultiRawDataOutput.csv')
	outdf = pd.DataFrame(columns = ['','Position384','Position96','Performance'])
	for index, row in df.iterrows() : 
		list_384 = row.Position384.split(' ')
		
		Alphabet_index = Alphabet_list.index(list_384[2][0])
		if Alphabet_index %2 ==0 :
			#find if another related plate no exist
			# 
			pair_plate = str(list_384[0])+' '+str(list_384[1])+' '+Alphabet_list[Alphabet_index+1]+str(list_384[2][1:]) 
			for i, row2 in df.iterrows() :
				if pair_plate == row2.Position384 :
					outdf = outdf.append(df.loc[index].copy(), ignore_index=True)
					print(i)
					break 
	#print (outdf) 
	return outdf
def multirawdataprocessfilerequirement() :
	messagebox.showinfo('提醒',str_multirawdataprocessfilerequirement)
	 
ttk.Label(window,  text="多功能384数据处理" ).grid(column=0, row=23) 
mv= IntVar()
ttk.Radiobutton(window, text ="倍数",variable=mv,value =1).grid(column=0,row= 24)
ttk.Radiobutton(window, text ="数值",variable=mv,value =2).grid(column=1,row= 24)
mv.set(1)

mthreshold384_entry=ttk.Entry(window,width = 8 )
mthreshold384_entry.grid(column=2, row=24) 
mthreshold384_entry.insert(0,"5")

mv2= IntVar()
ttk.Radiobutton(window, text ="大于",variable=mv2,value =1).grid(column=0,row= 25)
ttk.Radiobutton(window, text ="小于等于",variable=mv2,value =2).grid(column=1,row= 25)
mv2.set(1)

mv3= IntVar()
ttk.Radiobutton(window, text ="4x96",variable=mv3,value =1).grid(column=0,row= 26)
ttk.Radiobutton(window, text ="    2x96",variable=mv3,value =2).grid(column=1,row= 26)
mv3.set(1)


ttk.Label(window, text="文件名：").grid(column=0, row=27) 
lbmultirawdata =ttk.Label(window, text=" ")
lbmultirawdata.grid(column=1, row=27)

ttk.Button(window, text="文件说明", command=multirawdataprocessfilerequirement).grid(column=0, row=28)
ttk.Button(window, text="选择文件", command=chosemultirawdatafile).grid(column=1, row=28)
ttk.Button(window, text="处理文件", command=multirawdataprocess).grid(column=0, row=29)
########################################处理原始数据###########################################################################
dir_path = ''
save_excel_file = 'rawdata-collection.xlsx'

def rawfilerequirement() :
	messagebox.showinfo('提醒',str_extractrawdataprocessfilerequirement)

def choosedirectory():
	global dir_path
	dir_path = filedialog.askdirectory() 	
	str_list= dir_path.split('/')	
	lbexatrctrawdata.configure(text=str_list[-1])
def gettablenum(filepath):
	i= filepath.find('-')	
	return filepath[i+1:i+3]

def rawfileprocess():
	wb_write = Workbook() 
	for path in os.listdir(dir_path):
		if os.path.isfile(os.path.join(dir_path,path)):
			if filetypeRB.get() == 1:
				wb_read=load_workbook(os.path.join(dir_path,path))
			else : 
				wb_read = pd.read_csv(os.path.join(dir_path,path),header=None) 

			tabnum = gettablenum(path)
			ws_temp = wb_write.create_sheet(tabnum)
			line = int(line_entry.get())
			for r in range(17):
				for c in range(25):
					if filetypeRB.get() == 1:
						ws_temp.cell(row=r+1, column=c+1).value = wb_read.active.cell(row=r+line, column=c+1).value
					else : 
						ws_temp.cell(row=r+1, column=c+1).value = wb_read.iloc[r+line-1, c]
			ws_temp['A1'] = tabnum
	del wb_write['Sheet']
	wb_write.save(save_excel_file)
	messagebox.showinfo('提醒',"处理完成")

ttk.Label(window,  text="提取原始数据" ).grid(column=0, row=50) 
ttk.Label(window, text="目录名：").grid(column=0, row=51) 
lbexatrctrawdata =ttk.Label(window, text=" ")
lbexatrctrawdata.grid(column=1, row=51)

ttk.Button(window, text="模块说明", command=rawfilerequirement).grid(column=0, row=52)
ttk.Button(window, text="选择目录", command=choosedirectory).grid(column=1, row=52)
ttk.Label(window,justify="left", text="首行数：" ).grid(column=2, row=52)
line_entry=ttk.Entry(window,width = 8 )
line_entry.grid(column=3, row=52) 
line_entry.insert(0,"36")
filetypeRB= IntVar()
ttk.Radiobutton(window, text ="xlsx",variable=filetypeRB,value =1).grid(column=0,row= 53)
ttk.Radiobutton(window, text ="csv",variable=filetypeRB,value =2).grid(column=1,row= 53)
filetypeRB.set(1)
ttk.Button(window, text="处理文件", command=rawfileprocess).grid(column=0, row=54)
window.mainloop() 